# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.

import logging
import xlrd
import xlwt
import sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from odoo import _
from odoo import api, fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)


class product_kks_stock_wizard(models.Model):

    _name = "product.kks.stock.wizard"
    _description = "Product KKS stock Wizard"

    file_name = fields.Char("Nom du fichier", size=256)
    file = fields.Binary("Fichier", required=False)

    @api.multi
    def update_stock(self):
        reload(sys)
        sys.setdefaultencoding("utf-8")
        obj_stock = self.env["product.kks.stock"]
        results = self.env["product.image.directory"].search(
            [("type", "=", "reporting")]
        )
        for result in results:
            directory = result.name

        for data in self:
            f = open(directory + data.file_name, "wb")
            f.write(data.file.decode("base64"))
            f.close()
            extension = data.file_name.split(".")[1]
            if extension == "csv":
                rownum = 0
                lines = (data.file.decode("base64")).split("\n")
                for line in lines[:-1]:
                    if rownum == 0:
                        header = line.split(";")
                    else:
                        magasin = stock = ""
                        data = line.split(";")
                        magasin = data[0]
                        stock = data[1]
                        info_id = (
                            self.env["product.info"]
                            .search([("name", "=", "Stock")], limit=1)
                            .id
                        )
                        magasin_id = (
                            self.env["product.magasin"]
                            .search([("code", "=", magasin)], limit=1)
                            .id
                        )
                        if magasin_id and info_id:
                            rslt = obj_stock.search(
                                [
                                    ("magasin_id", "=", magasin_id),
                                    ("info_id", "=", info_id),
                                ],
                                limit=1,
                            )
                            if rslt:
                                rslt.write({"value": stock})
                            else:
                                obj_stock.create(
                                    {
                                        "magasin_id": magasin_id,
                                        "info_id": info_id,
                                        "value": stock,
                                    }
                                )

                    rownum += 1
            elif extension == "xls":
                wb = xlrd.open_workbook(directory + data.file_name)
                workbook = xlwt.Workbook(encoding="utf-8")
                style_yellow = xlwt.easyxf(
                    "font: bold 1, name Tahoma, height 160;"
                    "align: vertical center, horizontal center, wrap on;"
                    "borders: left thin, right thin, top thin, bottom thin;"
                )
                style_red = xlwt.easyxf(
                    "font: colour white, name Tahoma, height 160;"
                    "align: vertical center, horizontal center, wrap on;"
                    "borders: left thin, right thin, top thin, bottom thin;"
                    "pattern: pattern solid, pattern_fore_colour red, pattern_back_colour red"
                )
                style_green = xlwt.easyxf(
                    "font: colour white, name Tahoma, height 160;"
                    "align: vertical center, horizontal center, wrap on;"
                    "borders: left thin, right thin, top thin, bottom thin;"
                    "pattern: pattern solid, pattern_fore_colour blue, pattern_back_colour blue"
                )
                ws = workbook.add_sheet("Stock")
                ws.col(0).width = len("Mis à jour (Oui/Non)") * 256
                ws.col(2).width = len("Mis à jour (Oui/Non)") * 256
                sheet = wb.sheet_by_index(0)
                num_rows = sheet.nrows - 1
                # num_cells = sheet.ncols - 1
                num_cells = 1
                curr_row = -1
                while curr_row < num_rows:
                    curr_row += 1
                    curr_cell = -1
                    if curr_row == 0:
                        ws.write(curr_row, 0, "Magasin", style_yellow)
                        ws.write(curr_row, 1, "Stock", style_yellow)
                        ws.write(curr_row, 2, "Mis à jour (Oui/Non)", style_yellow)

                    else:
                        magasin = ""
                        stock = 0

                        while curr_cell < num_cells:
                            curr_cell += 1
                            cell_value = sheet.cell_value(curr_row, curr_cell)
                            if curr_cell == 0:
                                magasin = cell_value
                            if curr_cell == 1:
                                stock = cell_value
                                _logger.info(stock)
                                if isinstance(stock, (int, float)):
                                    stock = int(stock)
                                    _logger.info(stock)
                        _logger.info(stock)
                        info_id = (
                            self.env["product.info"]
                            .search([("name", "=", "Stock")], limit=1)
                            .id
                        )
                        magasin_id = (
                            self.env["product.magasin"]
                            .search([("code", "=", magasin)], limit=1)
                            .id
                        )
                        if magasin_id and info_id:
                            rslt = obj_stock.search(
                                [
                                    ("magasin_id", "=", magasin_id),
                                    ("info_id", "=", info_id),
                                ],
                                limit=1,
                            )
                            if rslt:
                                self._cr.execute(
                                    "update product_kks_stock set value='"
                                    + str(stock)
                                    + "' \
                                            where magasin_id="
                                    + str(magasin_id)
                                    + "\
                                            and info_id="
                                    + str(info_id)
                                )
                                # rslt.write({'value' : stock})
                            else:
                                self._cr.execute(
                                    "insert into product_kks_stock(magasin_id,info_id,value)\
                                         values ('"
                                    + str(magasin_id)
                                    + "',"
                                    + str(info_id)
                                    + ","
                                    + str(stock)
                                    + ")"
                                )
                                # obj_stock.create({'magasin_id' : magasin_id,'info_id' : info_id, 'value' : stock})
                            ws.write(curr_row, 0, magasin, style_green)
                            ws.write(curr_row, 1, stock, style_green)
                            ws.write(curr_row, 2, "Oui", style_green)
                        else:
                            ws.write(curr_row, 0, magasin, style_red)
                            ws.write(curr_row, 1, stock, style_red)
                            ws.write(curr_row, 2, "Non", style_red)

                workbook.save(directory + data.file_name)
            elif extension == "xlsx":
                workbook = load_workbook(directory + data.file_name)
                first_sheet = workbook.active
                j = 0
                for row in first_sheet.iter_rows():
                    j = j + 1
                    if j == 1:
                        first_sheet["C" + str(j)] = "Mis à jour (Oui/Non)"
                    else:
                        magasin = ""
                        stock = 0
                        i = 0
                        for cell in row:
                            i = i + 1
                            if i == 1:
                                magasin = cell.value
                            if i == 2:
                                stock = cell.value

                        info_id = (
                            self.env["product.info"]
                            .search([("name", "=", "Stock")], limit=1)
                            .id
                        )
                        magasin_id = (
                            self.env["product.magasin"]
                            .search([("code", "=", magasin)], limit=1)
                            .id
                        )
                        if magasin_id and info_id:
                            rslt = obj_stock.search(
                                [
                                    ("magasin_id", "=", magasin_id),
                                    ("info_id", "=", info_id),
                                ],
                                limit=1,
                            )
                            if rslt:
                                self._cr.execute(
                                    "update product_kks_stock set value='"
                                    + str(stock)
                                    + "' \
                                            where magasin_id="
                                    + str(magasin_id)
                                    + "\
                                            and info_id="
                                    + str(info_id)
                                )
                                # rslt.write({'value' : stock})
                            else:
                                self._cr.execute(
                                    "insert into product_kks_stock(magasin_id,info_id,value)\
                                         values ('"
                                    + str(magasin_id)
                                    + "',"
                                    + str(info_id)
                                    + ","
                                    + str(stock)
                                    + ")"
                                )
                                # obj_stock.create({'magasin_id' : magasin_id,'info_id' : info_id, 'value' : stock})

                            first_sheet["C" + str(j)] = "Oui"
                            first_sheet["A" + str(j)].fill = PatternFill(
                                start_color="00FF00",
                                end_color="00FF00",
                                fill_type="solid",
                            )
                            first_sheet["B" + str(j)].fill = PatternFill(
                                start_color="00FF00",
                                end_color="00FF00",
                                fill_type="solid",
                            )
                            first_sheet["C" + str(j)].fill = PatternFill(
                                start_color="00FF00",
                                end_color="00FF00",
                                fill_type="solid",
                            )
                        else:
                            first_sheet["C" + str(j)] = "Non"
                            first_sheet["A" + str(j)].fill = PatternFill(
                                start_color="FF0000",
                                end_color="FF0000",
                                fill_type="solid",
                            )
                            first_sheet["B" + str(j)].fill = PatternFill(
                                start_color="FF0000",
                                end_color="FF0000",
                                fill_type="solid",
                            )
                            first_sheet["C" + str(j)].fill = PatternFill(
                                start_color="FF0000",
                                end_color="FF0000",
                                fill_type="solid",
                            )
                workbook.save(directory + data.file_name)
            else:
                raise UserError(_("Seuls les fichiers CSV, XLS et XLSX sont autorisés"))
        url = "/web/static/reporting/" + self.file_name
        return {
            "type": "ir.actions.act_url",
            "url": url,
            "target": "self",
        }


class product_magasin_stock_wizard(models.Model):

    _name = "product.magasin.stock.wizard"
    _description = "Product Magasin Stock Wizard"

    magasin_id = fields.Many2one("product.magasin", "Magasin")
    info_ids = fields.Many2many(
        "product.info", "rel_magasin_info", "magasin_id", "info_id", "Infos"
    )

    def action_execute(self):
        for obj in self.info_ids:
            self.env["product.kks.stock"].create(
                {
                    "magasin_id": self.magasin_id.id,
                    "info_id": obj.id,
                }
            )
