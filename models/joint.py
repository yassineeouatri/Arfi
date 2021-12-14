# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.
import logging
import math
import time
import datetime
from odoo import api, fields, models
import sys
from openpyxl import Workbook, load_workbook
# If you need to get the column letter, also import this
from openpyxl.utils import get_column_letter
_logger = logging.getLogger(__name__)


class product_joint_torque(models.Model):
    _name = "product.joint.torque"

    def _default_classe_id(self):
        return self.env['product.joint.classe'].search([('name', '=', 'Class 150')], limit=1).id

    joint_id = fields.Many2one("product.joint.joint", "Joint")
    joint = fields.Char('Joint')
    fluide_id = fields.Many2one("product.joint.fluide", "Fluide")
    quality_id = fields.Many2one("product.joint.quality", "Qualité")
    norme_id = fields.Many2one("product.joint.norme", "Norme")
    type_id = fields.Many2one("product.joint.type", "Face")
    dn_id = fields.Many2one("product.joint.dn", "DN")
    emboitement_line_id = fields.Many2one("product.joint.emboitement.line", "DN")
    spiralise_id = fields.Many2one("product.joint.spiralise3", "DN")
    pn_id = fields.Many2one("product.joint.pn", "PN")
    classe_id = fields.Many2one('product.joint.classe', 'Classe', default=_default_classe_id)
    ss_anneau = fields.Selection(
        [('double' , 'Emboitement double'),
        ('double etroit', 'Emboitement double étroit'),
        ('double large', 'Emboitement double large'),
        ], 'Sans anneaux'
    )
    avec_anneau_int_ext = fields.Selection(
        [
        ('type1', 'Norme ASME B16-20 et Norme ASME B16.5'),
        ('type2', 'NF E 29-203 et/ou NF EN 1092-1'),
        ('type3', 'NF EN 12560-2M* et prEN 1759-1')], 'Avec anneaux ext ou int+ext'
    )
    avec_anneau_int = fields.Selection(
        [('simple' , 'Emboitement simple'),
        ('simple large', 'Emboitement simple large')], 'Avec anneaux int'
    )
    d_int_contact = fields.Float('Diamètre de contact (mm)')
    d_int_decoupe = fields.Float('diamètre de découpe (mm)')
    d_ext_contact = fields.Float('Diamètre de contact (mm)')
    d_ext_decoupe = fields.Float('diamètre de découpe (mm)')
    epaisseur = fields.Float('Epaisseur du joint (mm)', default=2)
    q = fields.Float('Q (MPa )', default=75)
    pression = fields.Integer('Pression de design Pd ( bar )', default=15)
    temperature = fields.Integer(' Temperature fuide  Tw (℃)', default=50)
    nb_boulon = fields.Integer(' Qty of Bolt n')
    d_boulon = fields.Float('Outer Dianmenter of Bolt d (mm)')
    d_trou = fields.Float('Option pleins trous : diamètre des trous', default=36)
    friction = fields.Float('Friction Coefficient with lubrication',default=0.15)
    contrainte = fields.Char('Contrainte', readonly="1")
    comptabilite = fields.Char('Comptabilité', readonly="1")
    d2_boulon = fields.Float(
        compute="_compute_d2_boulon", string="De (mm)", store=True, readonly=True
    )
    pas = fields.Float(
        compute="_compute_d2_boulon", string='pas p (mm)', store=True, readonly=True
    )
    coef_serrage = fields.Float('incertitude de serrage', default=1.2)
    assise = fields.Float(
        compute="_compute_assise", string="assise mini : (MPa)", store=True, readonly=True
    )
    surface = fields.Float(
        compute="_compute_surface", string=" Surface Area of Sealing Face S=π/4(Φ12-Φ22) (mm2 )", store=True, readonly=True
    )
    surface_n = fields.Char(
        compute="_compute_surface", string=" Surface Area of Sealing Face S=π/4(Φ12-Φ22) (mm2 )", store=True, readonly=True
    )
    fj = fields.Float(
        compute="_compute_fj", string="Fj= Q x S  (N)", store=True, readonly=True
    )
    fj_n = fields.Char(
        compute="_compute_fj", string="Fj= Q x S  (N)", store=True, readonly=True
    )
    ff = fields.Float(
        compute="_compute_ff", string=" Ff=Pd(π/4)Φm2 (N)", store=True, readonly=True
    )
    ff_n = fields.Char(
        compute="_compute_ff", string=" Ff=Pd(π/4)Φm2 (N)", store=True, readonly=True
    )
    ft = fields.Float(
        compute="_compute_ft", string="( Ft=Fj+Ff)  x incertitude serrage   (N)", store=True, readonly=True
    )
    ft_n = fields.Char(
        compute="_compute_ft", string="( Ft=Fj+Ff)  x incertitude serrage   (N)", store=True, readonly=True
    )
    fb = fields.Float(
        compute="_compute_fb", string="Fb=Ft/n (N)", store=True, readonly=True
    )
    fb_n = fields.Char(
        compute="_compute_fb", string="Fb=Ft/n (N)", store=True, readonly=True
    )
    mn = fields.Float(
        compute="_compute_mn", string="Mn =Fb[0.16xp+μ(0.583d2+rm)]  (N.m)", store=True, readonly=True
    )
    mn2 = fields.Float(
        compute="_compute_mn", string="Mn : sans coefficient de serrage ", store=True, readonly=True
    )
    d1 = fields.Float(
        compute="_compute_d", string="Inner diameter of nut D1=d-1.0825d (mm)", store=True, readonly=True
    )
    d2 = fields.Float(
        compute="_compute_d", string=" d2= d-0.6495p (mm)", store=True, readonly=True
    )
    d3 = fields.Float(
        compute="_compute_d", string="d3 =d-1.2268p  (mm)", store=True, readonly=True
    )
    d1_n = fields.Char(
        compute="_compute_d", string="Inner diameter of nut D1=d-1.0825d (mm)", store=True, readonly=True
    )
    d2_n = fields.Char(
        compute="_compute_d", string=" d2= d-0.6495p (mm)", store=True, readonly=True
    )
    d3_n = fields.Char(
        compute="_compute_d", string="d3 =d-1.2268p  (mm)", store=True, readonly=True
    )
    section_bolt = fields.Float(
        compute="_compute_section_bolt", string="Section Area of Bolt s= (π/4)d32  (mm2 )", store=True, readonly=True
    )
    rm_bolt = fields.Float(
        compute="_compute_rm_bolt", string="rm= (1/3)(De3-D13)/(De2-D12) (mm)", store=True, readonly=True
    )
    contrainte_joint = fields.Char(
        compute="_compute_contrainte_joint", string="contrainte sur le joint au montage :", store=True, readonly=True
    )
    contrainte_joint2 = fields.Char(
        compute="_compute_contrainte_joint2", string="Contrainte maxi admissible sur le joint (MPa) :", store=True, readonly=True
    )
    e = fields.Integer('E (MPa)')
    l = fields.Integer('L (mm)')
    delta = fields.Float(
        compute="_compute_delta", string="'deltaL (mm)'", store=True, readonly=True
    )
    contrainte_boulon = fields.Integer(
        compute="_compute_contrainte_boulon", string="contrainte boulon calculée :", store=True, readonly=True
    )
    l = fields.Integer('L (mm)')
    max_boulon = fields.Integer("Max bolt")
    pourcentage_boulon = fields.Char(
        compute="_compute_pourcentage_boulon", string="%", store=True, readonly=True
    )
    class_boulon = fields.Selection([
        ('class1', 'classe 4.6'),
        ('class2', 'classe 4.8'),
        ('class3', 'classe 5.6'),
        ('class4', 'classe 5.8'),
        ('class5', 'classe 6.8'),
        ('class6', 'classe 8.8'),
        ('class7', 'classe 9.8/B7'),
        ('class8', 'classe 10.9'),
        ('class9', 'classe 12.9')
    ], 'Classe')

    @api.one
    @api.depends("d_boulon")
    def _compute_d2_boulon(self):
        if self.d_boulon:
            if self.d_boulon == -1:
                self.d2_boulon = -1
                self.pas = -1
            else:
                if self.d_boulon == 6:
                    self.pas = 1
                elif self.d_boulon == 8:
                    self.pas = 1.25
                elif self.d_boulon == 10:
                    self.pas = 1.5
                elif self.d_boulon == 12:
                    self.pas = 1.75
                elif self.d_boulon >= 14 and self.d_boulon <= 16:
                    self.pas = 2
                elif self.d_boulon >= 18 and self.d_boulon <= 22:
                    self.pas = 2.5
                elif self.d_boulon >= 24 and self.d_boulon <= 27:
                    self.pas = 3
                elif self.d_boulon >= 30 and self.d_boulon <= 33:
                    self.pas = 3.5
                elif self.d_boulon >= 36 and self.d_boulon <= 39:
                    self.pas = 4
                elif self.d_boulon >= 42 and self.d_boulon <= 45:
                    self.pas = 4.5
                elif self.d_boulon >= 48 and self.d_boulon <= 52:
                    self.pas = 5
                elif self.d_boulon >= 56 and self.d_boulon <= 60:
                    self.pas = 5.5
                elif self.d_boulon >= 64 and self.d_boulon <= 90:
                    self.pas = 6
                else:
                    self.pas = -1
                self.d2_boulon = self.d_boulon + 2            

    @api.one
    @api.depends("pression")
    def _compute_assise(self):
        if self.pression:
            self.assise = (250 + 1.5 * self.pression) / 10
        else:
            self.assise = 0      

    @api.one
    @api.depends("d_int_contact","d_ext_contact", "nb_boulon", "d_trou")
    def _compute_surface(self):
        if self.nb_boulon!=-1:
            self.surface = (math.pi * (self.d_ext_contact * self.d_ext_contact - self.d_int_contact * self.d_int_contact)/ 4)  -(self.nb_boulon * math.pi * self.d_trou**2/4)
            self.surface_n = int(round((math.pi * (self.d_ext_contact * self.d_ext_contact - self.d_int_contact * self.d_int_contact)/ 4)  -(self.nb_boulon * math.pi * self.d_trou**2/4)))
        else:
            self.surface = -1
            self.surface_n = -1
    @api.one
    @api.depends("surface", "q")
    def _compute_fj(self):
        if self.surface != -1:
            self.fj = self.surface * self.q
            self.fj_n = int(abs(round(self.surface * self.q)))
        else :
            self.fj = -1
            self.fj_n = -1

    @api.one
    @api.depends("d_int_contact","d_ext_contact", "pression")
    def _compute_ff(self):
        self.ff = (math.pi/4)*((self.d_int_contact+self.d_ext_contact)/2)**2*self.pression*0.1
        self.ff_n = int(abs(round((math.pi/4)*((self.d_int_contact+self.d_ext_contact)/2)**2*self.pression*0.1)))

    @api.one
    @api.depends("fj","ff", "coef_serrage")
    def _compute_ft(self):
        if self.fj != -1:
            self.ft = (self.fj + self.ff) * self.coef_serrage
            self.ft_n = int(abs(round((self.fj + self.ff) * self.coef_serrage)))
        else:
            self.ft = -1
            self.ft_n = -1

    @api.one
    @api.depends("ft", "nb_boulon")
    def _compute_fb(self):
        if self.nb_boulon != -1:
            if self.nb_boulon == 0:
                self.fb = 0
                self.fb_n = 0
            else:
                self.fb = self.ft / self.nb_boulon
                self.fb_n = int(abs(round(self.ft / self.nb_boulon)))
        else:
            self.fb = -1
            self.fb_n = -1
    
    @api.one
    @api.depends("d_boulon", "pas")
    def _compute_d(self):
        if self.d_boulon!=-1:
            self.d1 = self.d_boulon - 1.0825 * self.pas
            self.d2 = self.d_boulon - 0.6495 * self.pas
            self.d3 = self.d_boulon - 1.2268 * self.pas
            self.d1_n = "{0:.1f}".format(self.d_boulon - 1.0825 * self.pas)
            self.d2_n = "{0:.2f}".format(self.d_boulon - 0.6495 * self.pas)
            self.d3_n = "{0:.2f}".format(self.d_boulon - 1.2268 * self.pas)
        else:
            self.d1 = -1
            self.d2 = -1
            self.d3 = -1
            self.d1_n = -1
            self.d2_n = -1
            self.d3_n = -1

    @api.one
    @api.depends("d3")
    def _compute_section_bolt(self):
        if self.d3!=-1:
            self.section_bolt = (3.14/4)*self.d3*self.d3
        else:
            self.section_bolt = -1

    @api.one
    @api.depends("d1", "d2_boulon")
    def _compute_rm_bolt(self):
        if self.d2_boulon!=-1:
            if (self.d2_boulon**2-self.d1**2) == 0:
                self.rm_bolt = 0
            else:
                self.rm_bolt = (1.0/3)*((self.d2_boulon**3-self.d1**3)*1.00/(self.d2_boulon**2-self.d1**2))
        else:
            self.rm_bolt = -1

    @api.one
    @api.depends("fb", "pas", "friction", "d2", "rm_bolt", "coef_serrage")
    def _compute_mn(self):
        if self.fb != -1:
            self.mn = round(self.fb*((0.16*self.pas+self.friction*(0.583*self.d2+self.rm_bolt)))/1000.0)
            self.mn2 = round((self.fb*((0.16*self.pas+self.friction*(0.583*self.d2+self.rm_bolt)))/1000.0)/self.coef_serrage)
        else:
            self.mn = -1
            self.mn2 = -1

    @api.one
    @api.depends("ft", "surface")
    def _compute_contrainte_joint(self):
        if self.surface != -1:
            if self.surface == 0:
                self.contrainte_joint = "0 MPa"
            else:
                self.contrainte_joint = "{} MPa".format(round(self.ft * 1.0/self.surface))
        else:
            self.contrainte_joint = -1

    @api.one
    @api.depends("d_ext_contact", "d_int_contact", "epaisseur", "temperature")
    def _compute_contrainte_joint2(self):
        self.contrainte_joint2 = "{} MPa".format(round((8.8*(self.d_ext_contact-self.d_int_contact)/2+40)*(-0.0093*self.epaisseur**2-0.0736*self.epaisseur+1.0815)*(-0.0007*self.temperature+1.0745)))
    
    @api.one
    @api.depends("e", "l", "q")
    def _compute_delta(self):
        if self.e == 0:
            self.delta = 0
        else:
            self.delta = self.q*self.l*1000/self.e

    @api.one
    @api.depends("fb", "section_bolt")
    def _compute_contrainte_boulon(self):
        if self.section_bolt != -1:
            if self.section_bolt == 0:
                self.contrainte_boulon = 0
            else:
                self.contrainte_boulon = round((self.fb*1.00/self.section_bolt))
        else:
            self.contrainte_boulon = -1
    
    @api.one
    @api.depends("contrainte_boulon", "max_boulon")
    def _compute_pourcentage_boulon(self):
        if self.contrainte_boulon != -1:
            if self.max_boulon == 0:
                self.pourcentage_boulon = 0
            else:
                self.pourcentage_boulon = "{:.0%} de re".format(self.contrainte_boulon * 1.00 / self.max_boulon)
        else:
            self.pourcentage_boulon = -1

    @api.onchange("fluide_id")
    def onchange_fluide_id(self):
        value = {}
        domain = {}
        value['comptabilite']=""
        quality_ids_list = []
        if self.fluide_id:
            cr = self.env.cr
            cr.execute(
                """select quality_id from rel_fluide_quality
                where fluide_id = {} """ .format(self.fluide_id.id)
            )
            for res in cr.fetchall():
                quality_ids_list.append(res[0])
            if self.fluide_id.concentration!=0:
                value['comptabilite']= "La compatibilité dépends du couple concentration/température contacter LATTY"
            else:
                value['comptabilite']=""
        else:
            value['comptabilite']=""
        domain = { 'quality_id': [ ('id', 'in', quality_ids_list) ], } 
        return { 'domain': domain, 'value': value }
    @api.onchange("quality_id")
    def onchange_quality_id(self):
        value = {}
        contmax = 0
        if self.quality_id:
            contrainte = self.env["product.joint.contrainte"].search([("quality_id", "=", self.quality_id.id),
                        ('epaisseur','=',self.epaisseur)])
            for obj in contrainte:
                contmax= obj.p
        if self.surface!=0:
            if self.contrainte_joint>contmax and self.epaisseur!=0 and self.joint_id:
                value['contrainte']="Contrainte= fonction (épaisseur,tempéture,type de joint) trop élevée contacter LATTY"
            else:
                value['contrainte']=""
        else:
            self.contrainte=""
        if self.epaisseur==0  and contmax==0:
            value['contrainte']="épaisseur non répertoriée contacter LATTY)"

        return {"value": value}
    @api.onchange("q")
    def onchange_q(self):
        value = {}
        contmax = 0
        if self.quality_id:
            contrainte = self.env["product.joint.contrainte"].search([("quality_id", "=", self.quality_id.id),
                        ('epaisseur','=',self.epaisseur)])
            for obj in contrainte:
                contmax= obj.p
        if self.surface!=0:
            if self.contrainte_joint>contmax and self.epaisseur!=0 and self.joint_id:
                value['contrainte']="Contrainte= fonction (épaisseur,tempéture,type de joint) trop élevée contacter LATTY"
            else:
                value['contrainte']=""
        else:
            self.contrainte=""
        if self.epaisseur==0  and contmax==0:
            value['contrainte']="épaisseur non répertoriée contacter LATTY)"

        return {"value": value}
    @api.onchange("temperature")
    def onchange_temperature(self):
        value = {}
        contmax = 0
        if self.quality_id:
            contrainte = self.env["product.joint.contrainte"].search([("quality_id", "=", self.quality_id.id),
                        ('epaisseur','=',self.epaisseur)])
            for obj in contrainte:
                contmax= obj.p
        
        if self.contrainte_joint>contmax and self.epaisseur!=0 and self.joint_id:
            value['contrainte']="Contrainte= fonction (épaisseur,tempéture,type de joint) trop élevée contacter LATTY"
        else:
            value['contrainte']=""
        return {"value": value}
    @api.onchange("class_boulon")
    def onchange_class_boulon(self):
        if self.class_boulon == 'class1':
            self.max_boulon = 240
        elif self.class_boulon == 'class2':
            self.max_boulon = 340
        elif self.class_boulon == 'class3':
            self.max_boulon = 300
        elif self.class_boulon == 'class4':
            self.max_boulon = 420
        elif self.class_boulon == 'class5':
            self.max_boulon = 480
        elif self.class_boulon == 'class6':
            self.max_boulon = 640
        elif self.class_boulon == 'class7':
            self.max_boulon = 720
        elif self.class_boulon == 'class8':
            self.max_boulon = 940
        elif self.class_boulon == 'class9':
            self.max_boulon = 1140
        else:
            self.max_boulon = 0

    @api.onchange("joint_id", "norme_id", "type_id", "dn_id", "pn_id", "ss_anneau", "avec_anneau_int", "avec_anneau_int_ext","emboitement_line_id", "pn_id", "classe_id")
    def onchange_joint_id(self):
        domain = {}
        value = {}
        if self.joint_id:
            if self.joint_id.alias == 'plat':
                type_ids_list = []
                dn_ids_list = []
                pn_ids_list = []
                if self.norme_id:
                    for obj in self.norme_id.type_ids:
                        type_ids_list.append(obj.id)
                    domain['type_id'] = [('id', 'in', type_ids_list)]
                    if self.type_id:
                        dimension = self.env["product.joint.dimension"].search([("norme_id", "=", self.norme_id.id)])
                        for obj in dimension:
                            dn_ids_list.append(obj.dn_id.id)
                        domain['dn_id'] = [('id', 'in', dn_ids_list)]
                        if self.dn_id:
                            dimension = self.env["product.joint.dimension"].search(
                                [("norme_id", "=", self.norme_id.id), ("type_id", "=", self.type_id.id), ("dn_id", "=", self.dn_id.id)])
                            for obj in dimension:
                                pn_ids_list.append(obj.pn_id.id)
                            domain['pn_id'] = [('id', 'in', pn_ids_list)]
                            if self.pn_id.id:
                                dimension = self.env["product.joint.dimension"].search(
                                            [("norme_id", "=", self.norme_id.id), ("type_id", "=", self.type_id.id)
                                            , ("dn_id", "=", self.dn_id.id), ("pn_id", "=", self.pn_id.id)])
                                if dimension:
                                    for obj in dimension:
                                        value['d_int_contact'] = obj.d1
                                        value['d_int_decoupe'] = obj.d1
                                        value['d_ext_contact'] = obj.d2
                                        value['d_ext_decoupe'] = obj.d2
                                        value['d_boulon'] = obj.d_boulon
                                        value['nb_boulon'] = obj.nb_boulon
                                else:
                                    value['d_int_contact'] = -1
                                    value['d_int_decoupe'] = -1
                                    value['d_ext_contact'] = -1
                                    value['d_ext_decoupe'] = -1
                                    value['d_boulon'] = -1
                                    value['nb_boulon'] = -1
                        else:
                            domain['pn_id'] = [('id', '=', -1)]
                    else:
                        domain['dn_id'] = [('id', '=', -1)]
                        domain['pn_id'] = [('id', '=', -1)]
                else:
                    domain['type_id'] = [('id', '=', -1)]
                    domain['dn_id'] = [('id', '=', -1)]
                    domain['pn_id'] = [('id', '=', -1)]
            elif self.joint_id.alias == 'ss_anneau':
                dn_ids_list = []
                pn_ids_list = []
                if self.ss_anneau:
                    cr = self.env.cr
                    cr.execute(
                        """select a.id from product_joint_emboitement_line a
                            inner join product_joint_emboitement b on a.emboitement_id=b.id
                            where b.name='{}' """ .format(self.ss_anneau)
                                        )
                    for res in cr.fetchall():
                        dn_ids_list.append(res[0])
                    cr.execute(
                        """select a.pn_id from rel_emboitement_pn a
                            inner join product_joint_emboitement b on a.emboitement_id=b.id
                            where b.name='{}' """ .format(self.ss_anneau)
                                        )
                    for res in cr.fetchall():
                        pn_ids_list.append(res[0])
                    if self.emboitement_line_id:
                        value['d_int_contact'] = self.emboitement_line_id.a2
                        value['d_int_decoupe'] = self.emboitement_line_id.a2
                        value['d_ext_contact'] = self.emboitement_line_id.a1
                        value['d_ext_decoupe'] = self.emboitement_line_id.a1
                        if self.pn_id:
                            boulon = self.env["product.joint.bolt"].search([ ("dn_id", "=", self.emboitement_line_id.dn_id.id), ("pn_id", "=", self.pn_id.id)])
                            if boulon:
                                for obj in boulon:
                                    self.nb_boulon = boulon.nb_boulon
                                    self.d_boulon = boulon.d_boulon
                            else:
                                self.nb_boulon = -1
                                self.d_boulon = -1
                domain['emboitement_line_id'] = [ ('id', 'in', dn_ids_list) ]
                domain['pn_id'] = [ ('id', 'in', pn_ids_list) ]
            elif self.joint_id.alias == 'avec_anneau_int':
                dn_ids_list = []
                pn_ids_list = []
                if self.avec_anneau_int:
                    cr = self.env.cr
                    cr.execute(
                        """select a.id from product_joint_emboitement_line a
                            inner join product_joint_emboitement b on a.emboitement_id=b.id
                            where b.name='{}' """ .format(self.avec_anneau_int)
                                        )
                    for res in cr.fetchall():
                        dn_ids_list.append(res[0])
                    cr.execute(
                        """select a.pn_id from rel_emboitement_pn a
                            inner join product_joint_emboitement b on a.emboitement_id=b.id
                            where b.name='{}' """ .format(self.avec_anneau_int)
                                        )
                    for res in cr.fetchall():
                        pn_ids_list.append(res[0])
                    if self.emboitement_line_id:
                        value['d_int_contact'] = self.emboitement_line_id.a2
                        value['d_int_decoupe'] = self.emboitement_line_id.d
                        value['d_ext_contact'] = self.emboitement_line_id.a1
                        value['d_ext_decoupe'] = self.emboitement_line_id.a1
                        if self.pn_id:
                            boulon = self.env["product.joint.bolt"].search([ ("dn_id", "=", self.emboitement_line_id.dn_id.id), ("pn_id", "=", self.pn_id.id)])
                            if boulon:
                                for obj in boulon:
                                    self.nb_boulon = boulon.nb_boulon
                                    self.d_boulon = boulon.d_boulon
                            else:
                                self.nb_boulon = -1
                                self.d_boulon = -1
                domain['emboitement_line_id'] = [ ('id', 'in', dn_ids_list) ]
                domain['pn_id'] = [ ('id', 'in', pn_ids_list) ]
            elif self.joint_id.alias == 'avec_anneau_int_ext':
                dn_ids_list = []
                pn_ids_list = []
                classe_ids_list = []
                if self.avec_anneau_int_ext:
                    cr = self.env.cr
                    if self.avec_anneau_int_ext == 'type1' or self.avec_anneau_int_ext == 'type3':
                        cr.execute("""select distinct dn_id from product_joint_spiralise3
                                    where type='{}' """ .format(self.avec_anneau_int_ext))
                        for res in cr.fetchall():
                            dn_ids_list.append(res[0])
                        if self.dn_id:
                            cr.execute("""select distinct classe_id from product_joint_spiralise3
                                        where type='{}' 
                                        and dn_id={}""" .format(self.avec_anneau_int_ext, self.dn_id.id) )
                            for res in cr.fetchall():
                                classe_ids_list.append(res[0])
                            if self.classe_id:
                                objs = self.env["product.joint.spiralise3"].search([("dn_id", "=", self.dn_id.id), 
                                            ("classe_id", "=", self.classe_id.id),
                                            ("type", "=", self.avec_anneau_int_ext)])
                                if objs:
                                    for obj in objs:
                                        value['d_int_contact'] = obj.a2
                                        value['d_int_decoupe'] = obj.d2
                                        value['d_ext_contact'] = obj.a1
                                        value['d_ext_decoupe'] = obj.d1
                                else:
                                    value['d_int_contact'] = -1
                                    value['d_int_decoupe'] = -1
                                    value['d_ext_contact'] = -1
                                    value['d_ext_decoupe'] = -1
                                boulon = self.env["product.joint.bolt"].search([ ("dn_id", "=", self.dn_id.id), ("classe_id", "=", self.classe_id.id)])
                                if boulon:
                                    for obj in boulon:
                                        self.nb_boulon = boulon.nb_boulon
                                        self.d_boulon = boulon.d_boulon
                                else:
                                    self.nb_boulon = -1
                                    self.d_boulon = -1
                    if self.avec_anneau_int_ext == 'type2':
                        cr.execute("""select distinct pn_id from product_joint_spiralise3
                            where type='{}' """ .format(self.avec_anneau_int_ext))
                        for res in cr.fetchall():
                            pn_ids_list.append(res[0])
                    
                        
                domain['dn_id'] =  [ ('id', 'in', dn_ids_list) ]
                domain['pn_id'] = [ ('id', 'in', pn_ids_list) ]
                domain['classe_id'] = [ ('id', 'in', classe_ids_list) ]
            value['joint'] = self.joint_id.alias
        return {"value": value, 'domain': domain}
    
    @api.multi
    def print_joint_excel(self):
        data = {}
        data["form"] = self.read()[0]
        return self._print_joint_excel(data)
    def get_return(self, fichier):
        url = "/web/static/reporting/" + fichier
        if url:
            return {
                "type": "ir.actions.act_url",
                "target": "new",
                "url": url,
                "nodestroy": True,
            }
        else:
            return True
    def _print_joint_excel(self, data):
        reload(sys)
        sys.setdefaultencoding("UTF8")
        results = self.env["product.image.directory"].search(
            [("type", "=", "reporting")]
        )
        for result in results:
            directory = result.name
        fichier = "joint.xlsx"
        wb = load_workbook(directory + fichier)
        ws = wb['FICHE DE SERRAGE']
        ws["G7"] = data['form']['pression']
        ws["G8"] = data['form']['temperature']
        ws["G12"] = data['form']['d_boulon']
        ws["G13"] = data['form']['pas']
        ws["G14"] = data['form']['nb_boulon']
        ws["C12"] = data['form']['d_int_contact']
        ws["C13"] = data['form']['d_ext_decoupe']
        ws["C14"] = data['form']['epaisseur']
        ws["F17"] = data['form']['mn']
        # Save file (if I don't provide a full file address, it goes to
        # the same folder as the script
        filename = "Joint_" + time.strftime("%H%M%S") + ".xlsx"
        wb.save(directory + filename)

        return self.get_return(filename)

class product_joint_dimension(models.Model):
    _name = "product.joint.dimension"

    norme_id = fields.Many2one("product.joint.norme", "Norme")
    type_id = fields.Many2one("product.joint.type", "Type")
    dn_id = fields.Many2one("product.joint.dn", "DN")
    pn_id = fields.Many2one("product.joint.pn", "PN")
    d1 = fields.Float('D1')
    d2 = fields.Float('D2')
    d = fields.Float('D portée')
    nb_boulon = fields.Integer('NB boulons')
    d_boulon = fields.Integer('Diam boulon')

class product_joint_emboitement(models.Model):
    _name = "product.joint.emboitement"

    name = fields.Char("Emboitement", required="1")
    pn_ids = fields.Many2many(
        "product.joint.pn", "rel_emboitement_pn", "emboitement_id", "pn_id", "PN"
    )
    emboitement_line_ids = fields.One2many(
        'product.joint.emboitement.line', 'emboitement_id', 'Lignes'
    )

class product_joint_emboitement_line(models.Model):
    _name = "product.joint.emboitement.line"

    @api.depends("nps", "dn_id")
    def compute_name(self):
        for obj in self:
            if obj.nps:
                obj.name = "DN {} NPS {}".format(obj.dn_id.name, obj.nps)
            else:
                obj.name = obj.dn_id.name

    name = fields.Char(compute="compute_name", store=True)
    emboitement_id = fields.Many2one('product.joint.emboitement', 'Emboitement')
    dn_id = fields.Many2one('product.joint.dn', 'DN')
    nps = fields.Char('NPS')
    d = fields.Float('d')
    a1 = fields.Float('a1')
    a2 = fields.Float('a2')
    epaisseur = fields.Float('Epaisseur')

class product_joint_spiralise3(models.Model):
    _name = "product.joint.spiralise3"

    @api.depends("nps", "dn_id")
    def compute_name(self):
        for obj in self:
            if obj.nps:
                obj.name = "DN {} NPS {}".format(obj.dn_id.name, obj.nps)
            else:
                obj.name = obj.dn_id.name
    
    name = fields.Char(compute="compute_name", store=True)
    type = fields.Selection(
        [('type1', 'Type1'),
        ('type2', 'Type2'),
        ('type3', 'Type3')]
    )
    classe_id = fields.Many2one('product.joint.classe', 'Classe')
    dn_id = fields.Many2one('product.joint.dn', 'DN')
    pn_id = fields.Many2one('product.joint.pn', 'PN')
    nps = fields.Char('NPS')
    d1 = fields.Float('D')
    d2 = fields.Float('d')
    a1 = fields.Float('a1')
    a2 = fields.Float('a2')

class product_joint_bolt(models.Model):
    _name = "product.joint.bolt"

    classe_id = fields.Many2one('product.joint.classe', 'Classe')
    dn_id = fields.Many2one('product.joint.dn', 'DN')
    pn_id = fields.Many2one('product.joint.pn', 'PN')
    nps = fields.Char('NPS')
    unc = fields.Char('UNC')
    nb_boulon = fields.Integer('NB boulons')
    d_boulon = fields.Integer('Diam boulon')

class product_joint_contrainte(models.Model):
    _name = "product.joint.contrainte"

    quality_id = fields.Many2one('product.joint.quality', 'Qualité')
    epaisseur = fields.Float('Epaisseur')
    f = fields.Integer('F')
    t = fields.Integer('T')
    p = fields.Float(
        compute="_compute_p", string="P", store=True, readonly=True
    )
    controle = fields.Integer(
        compute="_compute_p", string="Contrôle", store=True, readonly=True
    )
    limelastic = fields.Integer(
        compute="_compute_p", string="Limélastic", store=True, readonly=True
    )

    @api.one
    @api.depends("t", "f")
    def _compute_p(self):
        if self.quality_id.name == 'LATTYcarb 96':
            if self.epaisseur == 1:
                if self.t<=100:
                    self.p = "{0:.1f}".format(80)
                    self.limelastic = 1800
                else:
                    self.limelastic = int(-5.2*self.t+2320)
                    if self.t > 100 and self.t <=300:
                        self.p = "{0:.1f}".format(-0.1338*self.t+94.304)
                    else:
                        self.p = "{0:.1f}".format(-self.t+350)
                if self.t <= 350:
                    self.controle = int(self.p*self.f)
                else :
                    self.controle = 0
            elif self.epaisseur == 1.5:
                self.p ="{0:.1f}".format(-0.0016*self.t**2+0.15711*self.t+70)
                if self.t <= 245:
                    if self.t >=100:
                        self.controle = int(self.p*self.f)
                    else:
                        self.controle = int(70*self.f)
                else:
                    self.controle=0
                if self.t<=100:
                    self.limelastic=1650
                else:
                    self.limelastic=int(-6.6667*self.t+2166.7)
            elif self.epaisseur == 2:
                self.p = "{0:.1f}".format(-0.0016*self.t**2+0.15711*self.t+58.735)
                if self.t <=245:
                    if self.t >=100:
                        self.controle = int(self.p*self.f)
                    else:
                        self.controle = int(60*self.f)
                else:
                    self.controle=0
                if self.t<=100:
                    self.limelastic=1500
                else:
                    self.limelastic=int(-6.6667*self.t+2166.7)
            elif self.epaisseur == 3:
                self.p = "{0:.1f}".format(-0.0014*self.t**2+0.049*self.t+49.771)
                if self.t <=200:
                    if self.t >=100:
                        self.controle = int(self.p*self.f)
                    else:
                        self.controle = int(50*self.f)
                else:
                    self.controle=0
                if self.t<=50:
                    self.limelastic=1200
                else:
                    self.limelastic=int(-10*self.t+2500)
        elif self.quality_id.name == 'LATTYcarb 965':
            if self.epaisseur == 1:
                self.p = "{0:.1f}".format(-0.001*self.t**2+0.1742*self.t+90.919)
                if self.t <=400:
                    if self.t >=100:
                        self.controle = int(self.p*self.f)
                    else:
                        self.controle = int(105*self.f)
                else:
                    self.controle=0
                if self.t<=100:
                    self.limelastic=int(1800*1.1)
                else:
                    self.limelastic=int(-5.2*self.t+2320*1.1)
            elif self.epaisseur == 1.5:
                self.p = "{0:.1f}".format(-0.001*self.t**2+0.1742*self.t+80.919)
                if self.t <=350:
                    if self.t >=100:
                        self.controle = int(self.p*self.f)
                    else:
                        self.controle = int(86*self.f)
                else:
                    self.controle=0
                if self.t<=100:
                    self.limelastic=int(1650*1.1)
                else:
                    self.limelastic=int(-6.6667*self.t+2166.7*1.1)
            elif self.epaisseur == 2:
                self.p = "{0:.1f}".format(-0.001*self.t**2+0.147*self.t+71.303)
                if self.t <=350:
                    if self.t >=100:
                        self.controle = int(self.p*self.f)
                    else:
                        self.controle = int(82*self.f)
                else:
                    self.controle=0
                if self.t<=100:
                    self.limelastic=int(1500*1.1)
                else:
                    self.limelastic=int(-6.6667*self.t+2166.7*1.1)
            elif self.epaisseur == 3:
                self.p = "{0:.1f}".format(-0.0016*self.t**2+0.1571*self.t+60.735)
                if self.t <=250:
                    if self.t >=100:
                        self.controle = int(self.p*self.f)
                    else:
                        self.controle = int(60*self.f)
                else:
                    self.controle=0
                if self.t<=50:
                    self.limelastic=int(1200*1.1)
                else:
                    self.limelastic=int(-10*self.t+2500*1.1)
        elif self.quality_id.name == 'LATTYflon 95':
            if self.epaisseur == 1 or self.epaisseur == 1.5 or self.epaisseur == 2 or self.epaisseur == 3:
                self.p = "{0:.1f}".format(-0.5989*self.t+176.8)
                if self.t <=300:
                    if self.t >=50:
                        self.controle = int(self.p*self.f)
                    else:
                        self.controle = int(150*self.f)
                else:
                    self.controle=0
                if self.t<=50:
                    if self.epaisseur == 3:
                        self.limelastic=int(1500*0.8)
                    else:
                        self.limelastic=int(1500*0.9)
                else:
                    if self.epaisseur == 3:
                        self.limelastic=int(-5*self.t+1750*0.8)
                    else:
                        self.limelastic=int(-5*self.t+1750*0.9)
        elif self.quality_id.name == 'LATTYflon 95R':
            if self.epaisseur == 1.5 or self.epaisseur == 3:
                self.p = "{0:.1f}".format(-0.5989*self.t+176.8)
                if self.t <=300:
                    if self.t >=50:
                        self.controle = int(self.p*self.f)
                    else:
                        self.controle = int(150*self.f)
                else:
                    self.controle=0
                if self.t<=50:
                    if self.epaisseur == 1.5:
                        self.limelastic=int(1500*1.1)
                    else:
                        self.limelastic=int(1500*1)
                else:
                    if self.epaisseur == 1.5:
                        self.limelastic=int(-5*self.t+1750*1.1)
                    else:
                        self.limelastic=int(-5*self.t+1750*1)
        elif self.quality_id.name == 'LATTYgold 32':
            if self.epaisseur == 1:
                self.p = "{0:.1f}".format(-0.0013*self.t**2+0.1289*self.t+54.618)
                if self.t <=250:
                    if self.t >=50:
                        self.controle = int(self.p*self.f)
                    else:
                        self.controle = int(58*self.f)
                else:
                    self.controle=0
                if self.t<=100:
                    self.limelastic=int(1600*1)
                else:
                    self.limelastic=int(-5*self.t+1750*1)
            elif self.epaisseur == 1.5:
                self.p = "{0:.1f}".format(-0.001*self.t**2+0.0006*self.t+48.062)
                if self.t <=200:
                    if self.t >=50:
                        self.controle = int(self.p*self.f)
                    else:
                        self.controle = int(48*self.f)
                else:
                    self.controle=0
                if self.t<=50:
                    self.limelastic=int(1500*1)
                else:
                    self.limelastic=int(-6.6667*self.t+1833.3*1)
            elif self.epaisseur == 2:
                self.p = "{0:.1f}".format(-0.001*self.t**2+0.0006*self.t+41.062)
                if self.t <=200:
                    if self.t >=50:
                        self.controle = int(self.p*self.f)
                    else:
                        self.controle = int(40*self.f)
                else:
                    self.controle=0
                if self.t<=50:
                    self.limelastic=int(1500*1)
                else:
                    self.limelastic=int(-6.6667*self.t+1833.3*1)
            elif self.epaisseur == 3:
                self.p = "{0:.1f}".format(-0.3239*self.t+48.584)
                if self.t <=150:
                    if self.t >=50:
                        self.controle = int(self.p*self.f)
                    else:
                        self.controle = int(32*self.f)
                else:
                    self.controle=0
                if self.t<=50:
                    self.limelastic=int(1200*1)
                else:
                    self.limelastic=int(-8*self.t+1600*1)
        elif self.quality_id.name == 'LATTYgold 5 Acid':
            if self.epaisseur == 1:
                self.p = "{0:.1f}".format(-0.0007*self.t**2-0.1565*self.t+56.357)
                if self.t <=195:
                    self.controle = int(self.p*self.f)
                else:
                    self.controle=0
                if self.t<=100:
                    self.limelastic=int(1500*1)
                else:
                    self.limelastic=int(-10*self.t+2500*1)
            elif self.epaisseur == 1.5:
                self.p = "{0:.1f}".format(-0.0001*self.t**2-0.2044*self.t+47.136)
                if self.t <=180:
                    self.controle = int(self.p*self.f)
                else:
                    self.controle=0
                if self.t<=100:
                    self.limelastic=int(1500*1)
                else:
                    self.limelastic=int(-12.5*self.t+2750*1)
            elif self.epaisseur == 2:
                self.p = "{0:.1f}".format(-0.0001*self.t**2-0.2044*self.t+40.136)
                if self.t <=180:
                    self.controle = int(self.p*self.f)
                else:
                    self.controle=0
                if self.t<=100:
                    self.limelastic=int(1500*1)
                else:
                    self.limelastic=int(-12.5*self.t+2750*1)
            elif self.epaisseur == 3:
                self.p = "{0:.1f}".format(-0.000000000000000002*self.t**2-0.2*self.t+30)
                if self.t <=150:
                    self.controle = int(self.p*self.f)
                else:
                    self.controle=0
                if self.t<=50:
                    self.limelastic=int(1200*1)
                else:
                    self.limelastic=int(-8*self.t+1600*1)
        elif self.quality_id.name == 'LATTYgold 92':
            if self.epaisseur == 1:
                self.p = "{0:.1f}".format(-0.0011*self.t**2+0.1039*self.t+69.269)
                if self.t <=300:
                    if self.t >=100:
                        self.controle = int(self.p*self.f)
                    else:
                        self.controle = int(70*self.f)
                else:
                    self.controle=0
                if self.t<=100:
                    self.limelastic=int(1800*1)
                else:
                    self.limelastic=int(-4.7619*self.t+1928.6*1)
            elif self.epaisseur == 1.5:
                self.p = "{0:.1f}".format(-0.001*self.t**2+0.0184*self.t+60)
                if self.t <=250:
                    if self.t >=100:
                        self.controle = int(self.p*self.f)
                    else:
                        self.controle = int(60*self.f)
                else:
                    self.controle=0
                if self.t<=100:
                    self.limelastic=int(1500*1)
                else:
                    self.limelastic=int(-6.25*self.t+2062.5*1)
            elif self.epaisseur == 2:
                self.p = "{0:.1f}".format(-0.001*self.t**2+0.0184*self.t+51.924)
                if self.t <=250:
                    if self.t >=100:
                        self.controle = int(self.p*self.f)
                    else:
                        self.controle = int(50*self.f)
                else:
                    self.controle=0
                if self.t<=90:
                    self.limelastic=int(1500*1)
                else:
                    self.limelastic=int(-6.25*self.t+2062.5*1)
            elif self.epaisseur == 3:
                self.p = "{0:.1f}".format(-0.0008*self.t**2-0.0941*self.t+41.8744)
                if self.t <=200:
                    if self.t >=50:
                        self.controle = int(self.p*self.f)
                    else:
                        self.controle = int(40*self.f)
                else:
                    self.controle=0
                if self.t<=50:
                    self.limelastic=int(1200*1)
                else:
                    self.limelastic=int(-5.3333*self.t+1466.7*1)
        elif self.quality_id.name == 'LATTYgold 925':
            if self.epaisseur == 1:
                self.p = "{0:.1f}".format((-0.0011*self.t**2+0.1039*self.t+69.269)*1.1)
                if self.t <=300:
                    if self.t >=100:
                        self.controle = int(self.p*self.f*self.f*1.1)
                    else:
                        self.controle = int(70*self.f*self.f*1.1)
                else:
                    self.controle=0
                if self.t<=100:
                    self.limelastic=int(1800*1.1)
                else:
                    self.limelastic=int((-4.7619*self.t+1928.6)*1.1)
            elif self.epaisseur == 1.5:
                self.p = "{0:.1f}".format((-0.001*self.t**2+0.0184*self.t+60)*1.1)
                if self.t <=250:
                    if self.t >=100:
                        self.controle = int(self.p*self.f*self.f*1.1)
                    else:
                        self.controle = int(60*self.f*self.f*1.1)
                else:
                    self.controle=0
                if self.t<=100:
                    self.limelastic=int(1500*1.1)
                else:
                    self.limelastic=int((-6.25*self.t+2062.5*1)*1.1)
            elif self.epaisseur == 2:
                self.p = "{0:.1f}".format((-0.001*self.t**2+0.0184*self.t+51.924)*1.1)
                if self.t <=250:
                    if self.t >=100:
                        self.controle = int(self.p*self.f*self.f*1.1)
                    else:
                        self.controle = int(50*self.f*self.f*1.1)
                else:
                    self.controle=0
                if self.t<=90:
                    self.limelastic=int(1500*1.1)
                else:
                    self.limelastic=int((-6.25*self.t+2062.5)*1.1)
            elif self.epaisseur == 3:
                self.p = "{0:.1f}".format((-0.0008*self.t**2-0.0941*self.t+41.8744)*1.1)
                if self.t <=200:
                    if self.t >=50:
                        self.controle = int(self.p*self.f*self.f*1.1)
                    else:
                        self.controle = int(40*self.f*self.f*1.1)
                else:
                    self.controle=0
                if self.t<=50:
                    self.limelastic=int(1200*1.1)
                else:
                    self.limelastic=int((-5.3333*self.t+1466.7*1)*1.1)
        elif self.quality_id.name == 'LATTYgraf EFA':
            if self.epaisseur == 1:
                self.p = "{0:.1f}".format(-0.0005*self.t**2+0.0026*self.t+198.94)
                if self.t <=350:
                    self.controle = int(self.p*self.f)
                else:
                    if self.t>350 and self.t<=500:
                        self.controle = int(150*self.f)
                    else:
                        self.controle=0
                if self.t<=100:
                    self.limelastic=int(3000*1.1)
                else:
                    self.limelastic=int((-2.3333*self.t+2233.3)*1.1)
            elif self.epaisseur == 1.5:
                self.p = "{0:.1f}".format(-0.0005*self.t**2+0.0026*self.t+198.94)
                if self.t <=350:
                    self.controle = int(self.p*self.f)
                else:
                    if self.t>350 and self.t<=500:
                        self.controle = int(150*self.f)
                    else:
                        self.controle=0
                if self.t<=100:
                    self.limelastic=int(3000*1.1)
                else:
                    self.limelastic=int((-2.3333*self.t+2233.3)*1.1)
            elif self.epaisseur == 2:
                self.p = "{0:.1f}".format(-0.0004*self.t**2-0.028*self.t+180.52)
                if self.t <=350:
                    self.controle = int(self.p*self.f)
                else:
                    if self.t>350 and self.t<=500:
                        self.controle = int(130*self.f)
                    else:
                        self.controle=0
                if self.t<=100:
                    self.limelastic=int(3000*1)
                else:
                    self.limelastic=int((-2.3333*self.t+2233.3)*1)
            elif self.epaisseur == 3:
                self.p = "{0:.1f}".format(-0.0004*self.t**2-0.028*self.t+180.52)
                if self.t <=350:
                    self.controle = int(self.p*self.f)
                else:
                    if self.t>350 and self.t<=500:
                        self.controle = int(130*self.f)
                    else:
                        self.controle=0
                if self.t<=100:
                    self.limelastic=int(3000*1)
                else:
                    self.limelastic=int((-2.3333*self.t+2233.3)*1)
        elif self.quality_id.name == 'LATTYgraf EFI':
            if self.epaisseur == 1:
                self.p = "{0:.1f}".format(-0.0001*self.t**2-0.0857*self.t+122.99)
                if self.t <=350:
                    self.controle = int(self.p*self.f)
                else:
                    if self.t>350 and self.t<=500:
                        self.controle = int(70*self.f)
                    else:
                        self.controle=0
                if self.t<=100:
                    self.limelastic=int(3000)
                else:
                    self.limelastic=int((-1.3333*self.t+1133.3)*1)
            elif self.epaisseur == 1.5:
                self.p = "{0:.1f}".format(-0.0001*self.t**2-0.0857*self.t+122.99)
                if self.t <=350:
                    self.controle = int(self.p*self.f)
                else:
                    if self.t>350 and self.t<=500:
                        self.controle = int(70*self.f)
                    else:
                        self.controle=0
                if self.t<=100:
                    self.limelastic=int(3000)
                else:
                    self.limelastic=int((-1.3333*self.t+1133.3)*1)
            elif self.epaisseur == 2:
                self.p = "{0:.1f}".format(0.0002*self.t**2-0.2086*self.t+109.33)
                if self.t <=500:
                    self.controle = int(self.p*self.f)
                else:
                    self.controle=0
                if self.t<=100:
                    self.limelastic=int(3000*1.1)
                else:
                    self.limelastic=int((-1.3333*self.t+1133.3)*1.1)
            elif self.epaisseur == 3:
                self.p = "{0:.1f}".format(0.0002*self.t**2-0.2086*self.t+109.33)
                if self.t <=500:
                    self.controle = int(self.p*self.f)
                else:
                    self.controle=0
                if self.t<=100:
                    self.limelastic=int(3000*1.1)
                else:
                    self.limelastic=int((-1.3333*self.t+1133.3)*1.1)
        elif self.quality_id.name == 'LATTYgraf EFMC@':
            self.p = 1000
            self.controle = 1000
            self.limelastic=4000
        elif self.quality_id.name == 'LATTYflon 94L':
            self.p = "{0:.1f}".format(-0.5989*self.t+176.8)
            if self.t <=300:
                if self.t >= 50:
                    self.controle = int(self.p*self.f)
                else:
                    self.controle = int(150*self.f)
            else:
                self.controle=0
            if self.t<=50:
                self.limelastic=int(1500*0.9)
            else:
                self.limelastic=int((-5*self.t+1750)*0.9)
       
class product_joint_quality(models.Model):
    _name = "product.joint.quality"

    name = fields.Char("Qualité", size=256, required=True)
    code = fields.Integer("Code")

class product_joint_type(models.Model):
    _name = "product.joint.type"

    name = fields.Char("Type", size=256, required=True)
    code = fields.Integer("Code")

class product_joint_norme(models.Model):
    _name = "product.joint.norme"

    name = fields.Char("Norme", size=256, required=True)
    code = fields.Integer("Code")
    type_ids = fields.Many2many(
        "product.joint.type", "rel_norme_type", "norme_id", "type_id", "Types"
    )

class product_joint_classe(models.Model):
    _name = "product.joint.classe"

    name = fields.Char("Classe", size=256, required=True)

class product_joint_fluide(models.Model):
    _name = "product.joint.fluide"

    name = fields.Char("Fluide", size=256, required=True)
    concentration = fields.Integer("Concentration")
    quality_ids = fields.Many2many(
        "product.joint.quality", "rel_fluide_quality", "fluide_id", "quality_id", "Qualités"
    )

class product_joint_joint(models.Model):
    _name = "product.joint.joint"

    name = fields.Char("Joint", size=256, required=True)
    code = fields.Integer("Code")
    alias = fields.Char('alias')

class product_joint_face(models.Model):
    _name = "product.joint.face"

    name = fields.Char("Face", size=256, required=True)
    code = fields.Integer("Code")

class product_joint_dn(models.Model):
    _name = "product.joint.dn"
    _order = "name"

    name = fields.Integer("DN", size=256, required=True)
    code = fields.Integer("Code")

class product_joint_pn(models.Model):
    _name = "product.joint.pn"
    _order = "name"
    name = fields.Integer("PN", size=256, required=True)
    code = fields.Integer("Code")
